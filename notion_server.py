#!/usr/bin/env python3
# coding: utf-8
"""
案件登録サーバー
NotionのAPIへ直接登録します。
起動: python3 notion_server.py
"""

import sys, json, os, tempfile, datetime, uuid, hashlib, io
from http.server import HTTPServer, BaseHTTPRequestHandler
from urllib.request import urlopen, Request
from urllib.error import HTTPError, URLError
from urllib.parse import unquote
import json as jsonlib
import email, email.parser
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side
    OPENPYXL_OK = True
except ImportError:
    OPENPYXL_OK = False
    print("⚠️  openpyxlが未インストール。請求書xlsx生成を使うには: pip install openpyxl --break-system-packages")

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))

# .envファイルから環境変数を読み込む（Macローカル用）
_env_file = os.path.join(SCRIPT_DIR, ".env")
if os.path.exists(_env_file):
    with open(_env_file) as _f:
        for _line in _f:
            _line = _line.strip()
            if _line and not _line.startswith("#") and "=" in _line:
                _k, _v = _line.split("=", 1)
                os.environ.setdefault(_k.strip(), _v.strip().strip('"').strip("'"))

# ============================================================
# Notionトークンは環境変数 NOTION_TOKEN から読み込みます
# Railway: Variables タブで設定
# Mac ローカル: notion_server.py と同じフォルダに .env ファイルを作成し
#              NOTION_TOKEN=ntn_xxxxx と記載
# ============================================================
NOTION_TOKEN = os.environ.get("NOTION_TOKEN", "")

# ポート番号（Railway等クラウドは$PORT環境変数を使用）
PORT = int(os.environ.get("PORT", 8765))

# NotionのデータベースID（案件表）
CASE_DB_ID = "2e13288b-84b0-81fa-a8e8-fb22589ce378"

# お客様データベースID
CUSTOMER_DB_ID = "1513288b-84b0-8035-ae58-d410686d282d"

# 予定データベースID
SCHEDULE_DB_ID = "4a2a3ad54b5b41a6b138420ee5841ee3"

# 引き継ぎデータベースID（Notionで管理）
HANDOVER_DB_ID = "92c91778-a575-445b-80b8-f233a0c23261"

# 日次スケジュール保存ファイル（ローカルJSON）
DAILY_SCHEDULE_FILE = os.path.join(SCRIPT_DIR, "daily_schedules.json")

# ── 請求書システム（Notion DB）──────────────────────────
INVOICE_DB_ID = os.environ.get("INVOICE_DB_ID", "25f803a762fc4a4b8c876c8756b52b66")
WORKER_DB_ID  = os.environ.get("WORKER_DB_ID",  "17706159d9854ad4832a6e80a14e285f")

# ── 管理者パスワード認証 ────────────────────────────────
# パスワード変更方法：.env ファイルに ADMIN_PASSWORD=新しいパスワード と記載
# Railway Variables に ADMIN_PASSWORD=新しいパスワード でも可
# デフォルトパスワード: nts2026
ADMIN_PASSWORD = os.environ.get("ADMIN_PASSWORD", "")
# 認証トークン = sha256(パスワード + salt)
def _make_admin_token(pw):
    return hashlib.sha256((pw + "nts_admin_2026").encode()).hexdigest()
# デフォルトトークン = sha256("nts2026" + salt)。ADMIN_PASSWORDが設定されていれば上書き
_DEFAULT_ADMIN_TOKEN = "ec31d1b13ecbc5c119f79ae5ea02ce8a3f3265a354acd6063894c2b8bf863c8f"
ADMIN_TOKEN = _make_admin_token(ADMIN_PASSWORD) if ADMIN_PASSWORD else _DEFAULT_ADMIN_TOKEN

# PDF は一時保存（Railwayでは再デプロイで消えるが、請求書データはNotionに永続保存）
INVOICE_PDF_DIR = os.path.join(SCRIPT_DIR, "invoice_pdfs")
os.makedirs(INVOICE_PDF_DIR, exist_ok=True)

def parse_multipart(raw_body, content_type):
    """multipart/form-data をパースして fields, files を返す"""
    # Content-Typeからboundaryを抽出
    ct_str = str(content_type)
    boundary = None
    for part in ct_str.split(";"):
        part = part.strip()
        if part.startswith("boundary="):
            boundary = part[len("boundary="):].strip('"')
            break
    if not boundary:
        return {}, {}

    fields = {}
    files  = {}
    delimiter = ("--" + boundary).encode()
    end_delim  = ("--" + boundary + "--").encode()

    # bodyをパーツに分割
    for chunk in raw_body.split(delimiter):
        if not chunk or chunk.strip() in (b"", end_delim.lstrip(b"--" + boundary.encode())):
            continue
        if chunk.startswith(b"--"):   # 終端
            continue
        # ヘッダーとボディを分離（\r\n\r\n で区切る）
        if b"\r\n\r\n" not in chunk:
            continue
        header_part, body_part = chunk.split(b"\r\n\r\n", 1)
        # 末尾の \r\n を除去
        if body_part.endswith(b"\r\n"):
            body_part = body_part[:-2]

        headers = {}
        for line in header_part.split(b"\r\n"):
            if b":" in line:
                k, v = line.split(b":", 1)
                headers[k.strip().lower().decode()] = v.strip().decode("utf-8", errors="replace")

        disp = headers.get("content-disposition", "")
        if not disp:
            continue

        name, filename = "", ""
        for seg in disp.split(";"):
            seg = seg.strip()
            if seg.startswith("name="):
                name = seg[5:].strip('"')
            elif seg.startswith("filename="):
                filename = seg[9:].strip('"')

        if not name:
            continue
        if filename:
            files[name] = (filename, body_part)
        else:
            fields[name] = body_part.decode("utf-8", errors="replace")

    return fields, files

def load_daily_schedules():
    if os.path.exists(DAILY_SCHEDULE_FILE):
        try:
            with open(DAILY_SCHEDULE_FILE, "r", encoding="utf-8") as f:
                return json.load(f)
        except Exception:
            pass
    return {}

def save_daily_schedules(data):
    with open(DAILY_SCHEDULE_FILE, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)

# 顧客ページID
CUSTOMER_PAGES = {
    "201":       "3293288b-84b0-80dc-8adf-f83cb6b3b9a2",
    "301":       "2593288b-84b0-803b-9ddd-f188bb63d703",
    "302":       "2743288b-84b0-8096-b828-ed6983716d84",
    "307":       "2ed3288b-84b0-8017-9aef-fa72b1ff0ba4",
    "308":       "2f43288b-84b0-8005-b2d1-ec83b3d8d5b3",
    "311":       "31f3288b-84b0-80ba-aaae-c56eec3faa68",
    "315":       "31a3288b-84b0-805a-9dfd-e2285627b489",
    "316":       "e9377408-06cd-406f-a11f-cf972605078d",
    "3202-302":  "31f3288b-84b0-80db-8ed3-d56363e102ef",
    "3302-302":  "31a3288b-84b0-80c9-a868-d2198a02784c",
}

CUSTOMER_NAME_CACHE = {}   # {customer_no: customer_name}  キャッシュ

def get_customer_name_map():
    """全顧客のNo→名前マップを返す（初回だけNotionから取得してキャッシュ）"""
    if CUSTOMER_NAME_CACHE:
        return CUSTOMER_NAME_CACHE
    body = {"page_size": 100}
    cursor = None
    while True:
        if cursor: body["start_cursor"] = cursor
        result, _ = notion_request("POST", f"/databases/{CUSTOMER_DB_ID}/query", body)
        if not result: break
        for page in result.get("results", []):
            try:
                props = page["properties"]
                no   = (props["お客様No."]["rich_text"]   or [{}])[0].get("plain_text","").strip()
                name = (props["クライアント名"]["rich_text"] or [{}])[0].get("plain_text","").strip()
                if no and name:
                    CUSTOMER_NAME_CACHE[no] = name
            except Exception:
                pass
        if not result.get("has_more"): break
        cursor = result.get("next_cursor")
    print(f"  👥 顧客名キャッシュ構築: {len(CUSTOMER_NAME_CACHE)}件")
    return CUSTOMER_NAME_CACHE

NOTION_API = "https://api.notion.com/v1"
HEADERS = {
    "Authorization": f"Bearer {NOTION_TOKEN}",
    "Content-Type": "application/json",
    "Notion-Version": "2022-06-28",
}

# ========== 引き継ぎ（Notion DB）==========
def parse_handover_page(page):
    """NotionページをHandoverアイテムに変換"""
    try:
        props = page["properties"]
        text       = (props["内容"]["title"] or [{}])[0].get("plain_text","").strip()
        typ        = (props["種別"]["select"] or {}).get("name","todo")
        status     = (props["ステータス"]["select"] or {}).get("name","active")
        date       = (props["日付"]["date"] or {}).get("start","")
        start_date = (props["開始日"]["date"] or {}).get("start","")
        end_date   = (props["終了日"]["date"] or {}).get("start","")
        created_at = (props["作成日"]["date"] or {}).get("start","")
        item = {"id": page["id"], "type": typ, "text": text,
                "status": status, "created_at": created_at}
        if typ == "medium":
            item["start_date"] = start_date
            item["end_date"]   = end_date
        else:
            item["date"] = date
        return item
    except Exception as e:
        print(f"  ⚠️ parse_handover_page: {e}")
        return None

def get_active_handover(date_str):
    """指定日にアクティブな引き継ぎをNotionから取得"""
    body = {"filter": {"property": "ステータス", "select": {"equals": "active"}},
            "page_size": 100}
    result, _ = notion_request("POST", f"/databases/{HANDOVER_DB_ID}/query", body)
    if not result:
        return []
    items = []
    for page in result.get("results", []):
        item = parse_handover_page(page)
        if not item:
            continue
        t = item["type"]
        if t == "medium":
            if item.get("start_date","") <= date_str <= item.get("end_date",""):
                items.append(item)
        elif t == "todo":
            if item.get("date","") <= date_str:
                items.append(item)
        else:
            if item.get("date","") == date_str:
                items.append(item)
    return items

def get_done_handover():
    """完了済みの引き継ぎをNotionから取得"""
    body = {"filter": {"property": "ステータス", "select": {"equals": "done"}},
            "sorts": [{"timestamp": "last_edited_time", "direction": "descending"}],
            "page_size": 100}
    result, _ = notion_request("POST", f"/databases/{HANDOVER_DB_ID}/query", body)
    if not result:
        return []
    items = []
    for page in result.get("results", []):
        item = parse_handover_page(page)
        if item:
            items.append(item)
    return items

def notion_request(method, path, body=None):
    url = NOTION_API + path
    data = jsonlib.dumps(body).encode() if body else None
    req = Request(url, data=data, headers=HEADERS, method=method)
    try:
        with urlopen(req, timeout=15) as res:
            return jsonlib.loads(res.read()), None
    except HTTPError as e:
        err_body = e.read().decode(errors='replace')
        print(f"  ❌ Notion HTTP {e.code}: {err_body[:300]}")
        return None, f"Notion API {e.code}: {err_body[:200]}"
    except URLError as e:
        print(f"  ❌ Notion URLError: {e.reason}")
        return None, f"Notion接続エラー: {e.reason}"
    except Exception as e:
        print(f"  ❌ Notion例外: {e}")
        return None, f"例外: {e}"

# ========== 請求書 Notion ヘルパー ==========

def parse_notion_invoice_page(page):
    """Notion請求書ページをdictに変換"""
    try:
        props = page["properties"]
        def rt(key, default=""):
            lst = (props.get(key, {}).get("rich_text") or [])
            return lst[0].get("plain_text", default) if lst else default
        def title_prop():
            t = (props.get("タイトル", {}).get("title") or [])
            return t[0].get("plain_text", "") if t else ""
        def sel(key, default=""):
            return (props.get(key, {}).get("select") or {}).get("name", default)
        def num(key, default=0):
            v = props.get(key, {}).get("number")
            return v if v is not None else default
        def chk(key):
            return props.get(key, {}).get("checkbox", False)
        def date_prop(key):
            return (props.get(key, {}).get("date") or {}).get("start", "")
        return {
            "id":             title_prop(),
            "_notion_id":     page["id"],
            "received_at":    date_prop("受信日時"),
            "worker_name":    rt("ワーカー名"),
            "billing_month":  rt("請求月"),
            "amount":         num("金額"),
            "status":         sel("ステータス", "new"),
            "bank_name":      rt("金融機関名"),
            "branch":         rt("支店"),
            "account_type":   sel("口座種別", "普通"),
            "account_number": rt("口座番号"),
            "account_holder": rt("口座名義"),
            "remark":         rt("備考"),
            "has_pdf":        chk("PDFあり"),
            "worker_id":      rt("ワーカーID"),
        }
    except Exception as e:
        print(f"  ⚠️ parse_notion_invoice_page: {e}")
        return None

def invoice_to_notion_props(inv):
    """請求書dictをNotion propertiesに変換"""
    received_at = inv.get("received_at", "")
    date_str = received_at[:10] if received_at else datetime.date.today().isoformat()
    return {
        "タイトル":   {"title":     [{"text": {"content": inv["id"]}}]},
        "ワーカー名": {"rich_text": [{"text": {"content": inv.get("worker_name", "")}}]},
        "請求月":     {"rich_text": [{"text": {"content": inv.get("billing_month", "")}}]},
        "金額":       {"number":    inv.get("amount", 0)},
        "ステータス": {"select":    {"name": inv.get("status", "new")}},
        "金融機関名": {"rich_text": [{"text": {"content": inv.get("bank_name", "")}}]},
        "支店":       {"rich_text": [{"text": {"content": inv.get("branch", "")}}]},
        "口座種別":   {"select":    {"name": inv.get("account_type", "普通")}},
        "口座番号":   {"rich_text": [{"text": {"content": inv.get("account_number", "")}}]},
        "口座名義":   {"rich_text": [{"text": {"content": inv.get("account_holder", "")}}]},
        "備考":       {"rich_text": [{"text": {"content": inv.get("remark", "")}}]},
        "受信日時":   {"date":      {"start": date_str}},
        "PDFあり":    {"checkbox":  inv.get("has_pdf", False)},
        "ワーカーID": {"rich_text": [{"text": {"content": inv.get("worker_id", "") or ""}}]},
    }

def get_invoices_from_notion():
    """Notionから請求書一覧を取得（新しい順）"""
    if not INVOICE_DB_ID:
        return []
    results = []
    body = {
        "sorts": [{"timestamp": "created_time", "direction": "descending"}],
        "page_size": 100,
    }
    cursor = None
    while True:
        if cursor: body["start_cursor"] = cursor
        result, _ = notion_request("POST", f"/databases/{INVOICE_DB_ID}/query", body)
        if not result: break
        for page in result.get("results", []):
            inv = parse_notion_invoice_page(page)
            if inv: results.append(inv)
        if not result.get("has_more"): break
        cursor = result.get("next_cursor")
    return results

def find_invoice_notion_id(inv_id):
    """請求書IDからNotionページIDを検索"""
    if not INVOICE_DB_ID: return None
    body = {
        "filter": {"property": "タイトル", "title": {"equals": inv_id}},
        "page_size": 1,
    }
    result, _ = notion_request("POST", f"/databases/{INVOICE_DB_ID}/query", body)
    if result and result.get("results"):
        return result["results"][0]["id"]
    return None

# ========== ワーカー Notion ヘルパー ==========

def parse_notion_worker_page(page):
    """Notionワーカーページをdictに変換"""
    try:
        props = page["properties"]
        def rt(key, default=""):
            lst = (props.get(key, {}).get("rich_text") or [])
            return lst[0].get("plain_text", default) if lst else default
        def title_prop():
            t = (props.get("タイトル", {}).get("title") or [])
            return t[0].get("plain_text", "") if t else ""
        def sel(key, default=""):
            return (props.get(key, {}).get("select") or {}).get("name", default)
        return {
            "worker_id":  title_prop(),
            "_notion_id": page["id"],
            "name":       rt("名前"),
            "pw_hash":    rt("パスワードハッシュ"),
            "token":      rt("トークン"),
            "bankName":   rt("金融機関名"),
            "bankBranch": rt("支店"),
            "bankNo":     rt("口座番号"),
            "bankHolder": rt("口座名義"),
            "acctType":   sel("口座種別", "普通"),
            "note":       rt("備考メモ"),
        }
    except Exception as e:
        print(f"  ⚠️ parse_notion_worker_page: {e}")
        return None

def worker_to_notion_props(w):
    """ワーカーdictをNotion propertiesに変換"""
    return {
        "タイトル":           {"title":     [{"text": {"content": w.get("worker_id", "")}}]},
        "名前":               {"rich_text": [{"text": {"content": w.get("name", "")}}]},
        "パスワードハッシュ": {"rich_text": [{"text": {"content": w.get("pw_hash", "")}}]},
        "トークン":           {"rich_text": [{"text": {"content": w.get("token", "")}}]},
        "金融機関名":         {"rich_text": [{"text": {"content": w.get("bankName", "")}}]},
        "支店":               {"rich_text": [{"text": {"content": w.get("bankBranch", "")}}]},
        "口座番号":           {"rich_text": [{"text": {"content": w.get("bankNo", "")}}]},
        "口座名義":           {"rich_text": [{"text": {"content": w.get("bankHolder", "")}}]},
        "口座種別":           {"select":    {"name": w.get("acctType", "普通")}},
        "備考メモ":           {"rich_text": [{"text": {"content": w.get("note", "")}}]},
    }

def get_worker_by_id_notion(worker_id):
    """ワーカーIDでNotionからワーカーを検索"""
    if not WORKER_DB_ID: return None
    body = {
        "filter": {"property": "タイトル", "title": {"equals": worker_id}},
        "page_size": 1,
    }
    result, _ = notion_request("POST", f"/databases/{WORKER_DB_ID}/query", body)
    if result and result.get("results"):
        return parse_notion_worker_page(result["results"][0])
    return None

def get_worker_by_token_notion(token):
    """トークンでNotionからワーカーを検索"""
    if not WORKER_DB_ID or not token: return None
    body = {
        "filter": {"property": "トークン", "rich_text": {"equals": token}},
        "page_size": 1,
    }
    result, _ = notion_request("POST", f"/databases/{WORKER_DB_ID}/query", body)
    if result and result.get("results"):
        return parse_notion_worker_page(result["results"][0])
    return None

# ========== 顧客番号ユーティリティ ==========
def get_all_customer_nos():
    """お客様DBから全番号を取得"""
    nos = []
    cursor = None
    while True:
        body = {"page_size": 100}
        if cursor:
            body["start_cursor"] = cursor
        result, _ = notion_request("POST", f"/databases/{CUSTOMER_DB_ID}/query", body)
        if not result:
            break
        for page in result.get("results", []):
            try:
                rt = page["properties"]["お客様No."]["rich_text"]
                no = rt[0]["plain_text"].strip() if rt else ""
                if no:
                    nos.append(no)
            except Exception:
                pass
        if not result.get("has_more"):
            break
        cursor = result.get("next_cursor")
    return nos

def next_in_range(nos, lo, hi):
    used = set()
    for no in nos:
        try:
            n = int(no)
            if lo <= n <= hi:
                used.add(n)
        except Exception:
            pass
    for n in range(lo, hi + 2):
        if n not in used:
            return str(n)

class Handler(BaseHTTPRequestHandler):
    def log_message(self, fmt, *args):
        print(f"  {args[0]} {args[1]}")

    def send_json(self, code, obj):
        body = jsonlib.dumps(obj, ensure_ascii=False).encode()
        self.send_response(code)
        self.send_header("Content-Type", "application/json; charset=utf-8")
        self.send_header("Access-Control-Allow-Origin", "*")
        self.send_header("Access-Control-Allow-Methods", "GET, POST, OPTIONS")
        self.send_header("Access-Control-Allow-Headers", "Content-Type")
        self.end_headers()
        self.wfile.write(body)

    def do_OPTIONS(self):
        self.send_response(204)
        self.send_header("Access-Control-Allow-Origin", "*")
        self.send_header("Access-Control-Allow-Methods", "GET, POST, OPTIONS")
        self.send_header("Access-Control-Allow-Headers", "Content-Type")
        self.end_headers()

    def send_html(self, filename):
        filepath = os.path.join(SCRIPT_DIR, filename)
        if not os.path.exists(filepath):
            self.send_response(404)
            self.end_headers()
            return
        with open(filepath, "rb") as f:
            body = f.read()
        self.send_response(200)
        self.send_header("Content-Type", "text/html; charset=utf-8")
        self.send_header("Content-Length", len(body))
        self.end_headers()
        self.wfile.write(body)

    def do_GET(self):
        from urllib.parse import unquote
        path = unquote(self.path.split('?')[0])  # デコード＆クエリ除去

        if path in ("/", "/index.html"):
            self.send_html("index.html")
        elif path == "/案件登録ツール.html":
            self.send_html("案件登録ツール.html")
        elif path == "/顧客登録ツール.html":
            self.send_html("顧客登録ツール.html")
        elif path == "/請求書ツール.html":
            self.send_html("請求書ツール.html")
        elif path == "/顧客情報ツール.html":
            self.send_html("顧客情報ツール.html")
        elif path == "/引き継ぎ一覧.html":
            self.send_html("引き継ぎ一覧.html")
        elif path == "/引き継ぎ完了済み.html":
            self.send_html("引き継ぎ完了済み.html")
        elif path == "/日次スケジュール.html":
            self.send_html("日次スケジュール.html")
        elif path == "/案件履歴ツール.html":
            self.send_html("案件履歴ツール.html")
        elif path == "/請求書管理ツール.html":
            self.send_html("請求書管理ツール.html")
        elif path == "/ワーカー請求書.html":
            self.send_html("ワーカー請求書.html")
        elif path == "/api/invoices":
            invoices = get_invoices_from_notion()
            self.send_json(200, invoices)
        elif path.startswith("/api/invoice-pdf/"):
            inv_id = path.split("/")[-1]
            pdf_path = os.path.join(INVOICE_PDF_DIR, inv_id + ".pdf")
            if os.path.exists(pdf_path):
                with open(pdf_path, "rb") as f:
                    data = f.read()
                self.send_response(200)
                self.send_header("Content-Type", "application/pdf")
                self.send_header("Content-Length", str(len(data)))
                self.send_header("Access-Control-Allow-Origin", "*")
                self.end_headers()
                self.wfile.write(data)
            else:
                self.send_json(404, {"error": "PDF not found"})
        elif path == "/api/health":
            self.send_json(200, {
                "status": "ok",
                "message": "サーバー起動中",
                "invoice_db_set": bool(INVOICE_DB_ID),
                "worker_db_set":  bool(WORKER_DB_ID),
                "notion_token_set": bool(NOTION_TOKEN),
            })
        elif path == "/api/myip":
            ip = self.headers.get("X-Forwarded-For", self.client_address[0]).split(",")[0].strip()
            self.send_json(200, {"ip": ip})
        elif path == "/api/next-customer-no":
            nos = get_all_customer_nos()
            self.send_json(200, {
                "ok": True,
                "next_300s":    next_in_range(nos, 300, 399),
                "next_3000s":   next_in_range(nos, 3000, 3999),
                "next_regular": next_in_range(nos, 1, 99),
            })
        elif path == "/api/customers":
            self.handle_get_customers()
        elif path == "/api/customers-all":
            self.handle_get_all_customers()
        elif path.startswith("/api/invoice-data"):
            self.handle_get_invoice_data()
        elif path.startswith("/api/calendar/day"):
            self.handle_get_calendar_day()
        elif path.startswith("/api/calendar"):
            self.handle_get_calendar()
        elif path == "/api/handover/all":
            self.handle_get_handover_all()
        elif path == "/api/handover/done-list":
            self.handle_get_handover_done_list()
        elif path.startswith("/api/cases/list"):
            self.handle_get_cases_list()
        elif path.startswith("/api/daily-schedule/dates"):
            self.handle_get_schedule_dates()
        elif path.startswith("/api/daily-schedule"):
            self.handle_get_daily_schedule()
        elif path.startswith("/api/handover"):
            self.handle_get_handover()
        else:
            self.send_json(404, {"error": "Not found"})

    def do_DELETE(self):
        path = unquote(self.path.split('?')[0])
        if path.startswith("/api/invoice/") and path.count("/") == 3:
            inv_id = path.split("/")[-1]
            notion_page_id = find_invoice_notion_id(inv_id)
            if notion_page_id:
                notion_request("PATCH", f"/pages/{notion_page_id}", {"archived": True})
                print(f"  🗑️  請求書削除（アーカイブ）: {inv_id}")
            self.send_json(200, {"ok": True})
        else:
            self.send_json(404, {"error": "Not found"})

    def do_PATCH(self):
        path = unquote(self.path.split('?')[0])
        length = int(self.headers.get("Content-Length", 0))
        raw = self.rfile.read(length)
        try:
            data = jsonlib.loads(raw)
        except:
            self.send_json(400, {"error": "Invalid JSON"}); return
        # /api/invoice/{id}/status
        parts = path.strip("/").split("/")
        if len(parts) == 4 and parts[0] == "api" and parts[1] == "invoice" and parts[3] == "status":
            inv_id = parts[2]
            new_status = data.get("status", "new")
            notion_page_id = find_invoice_notion_id(inv_id)
            if notion_page_id:
                notion_request("PATCH", f"/pages/{notion_page_id}", {
                    "properties": {"ステータス": {"select": {"name": new_status}}}
                })
                print(f"  ✅ 請求書ステータス更新: {inv_id} → {new_status}")
            self.send_json(200, {"ok": True})
        else:
            self.send_json(404, {"error": "Not found"})

    def do_POST(self):
        path = unquote(self.path.split('?')[0])
        length = int(self.headers.get("Content-Length", 0))
        raw = self.rfile.read(length)
        content_type = self.headers.get("Content-Type", "")

        # ── multipart/form-data（請求書PDF送信）──────────
        if "multipart/form-data" in content_type:
            if path == "/api/submit-invoice":
                self.handle_submit_invoice(raw, content_type)
            else:
                self.send_json(404, {"error": "Not found"})
            return

        # ── JSON系エンドポイント ──────────────────────────
        try:
            data = jsonlib.loads(raw) if raw else {}
        except Exception:
            self.send_json(400, {"error": "Invalid JSON"})
            return

        if path == "/api/admin/login":
            pw = data.get("password", "")
            pw_token = _make_admin_token(pw)
            if ADMIN_TOKEN and pw_token == ADMIN_TOKEN:
                self.send_json(200, {"ok": True, "token": ADMIN_TOKEN})
            else:
                self.send_json(401, {"ok": False, "error": "パスワードが違います"})
            return
        elif path == "/api/admin/check":
            token = data.get("token", "")
            ok = bool(ADMIN_TOKEN) and token == ADMIN_TOKEN
            self.send_json(200, {"ok": ok})
            return
        if path == "/api/worker/login":
            self.handle_worker_login(data)
            return
        elif path == "/api/worker/register":
            self.handle_worker_register(data)
            return
        elif path.startswith("/api/register-worker-from-invoice/"):
            inv_id = path.split("/")[-1]
            self.handle_register_worker_from_invoice(inv_id, data)
            return
        if self.path == "/api/register":
            self.handle_register(data)
        elif self.path == "/api/update-customer":
            self.handle_update_customer(data)
        elif self.path == "/api/register-customer":
            self.handle_register_customer(data)
        elif self.path == "/api/record-invoice":
            self.handle_record_invoice(data)
        elif self.path == "/api/generate-invoice":
            self.handle_generate_invoice(data)
        elif self.path == "/api/calendar/add":
            self.handle_add_schedule(data)
        elif self.path == "/api/calendar/delete":
            self.handle_delete_schedule(data)
        elif self.path == "/api/customers/bulk-update":
            self.handle_bulk_update_customers(data)
        elif self.path == "/api/customers/bulk-archive":
            self.handle_bulk_archive_customers(data)
        elif self.path == "/api/handover/add":
            self.handle_handover_add(data)
        elif self.path == "/api/handover/done":
            self.handle_handover_done(data)
        elif self.path == "/api/handover/carry":
            self.handle_handover_carry(data)
        elif self.path == "/api/handover/extend":
            self.handle_handover_extend(data)
        elif self.path == "/api/handover/delete":
            self.handle_handover_delete(data)
        elif self.path == "/api/handover/update-date":
            self.handle_handover_update_date(data)
        elif self.path == "/api/handover/update-content":
            self.handle_handover_update_content(data)
        elif self.path == "/api/handover/restore":
            self.handle_handover_restore(data)
        elif self.path == "/api/daily-schedule/save":
            self.handle_save_daily_block(data)
        elif self.path == "/api/daily-schedule/delete":
            self.handle_delete_daily_block(data)
        elif self.path == "/api/notion-event/move":
            self.handle_move_notion_event(data)
        elif self.path == "/api/cases/update":
            self.handle_update_case(data)
        elif self.path == "/api/daily-schedule/gantt":
            self.handle_gantt_excel(data)
        elif self.path == "/api/invoice/template":
            self.handle_invoice_template(data)
        elif self.path == "/api/invoice/numbers":
            self.handle_invoice_numbers(data)
        else:
            self.send_json(404, {"error": "Not found"})

    # ── 請求書送信（multipart → Notion）──────────────────────────────
    def handle_submit_invoice(self, raw_body, content_type):
        try:
            fields, files = parse_multipart(raw_body, content_type)
        except Exception as e:
            self.send_json(400, {"ok": False, "error": f"パースエラー: {e}"}); return

        if not INVOICE_DB_ID:
            self.send_json(500, {"ok": False,
                "error": "INVOICE_DB_ID が設定されていません。Railway の Variables タブで設定してください"}); return

        inv_id = "inv-" + str(uuid.uuid4())[:8]
        now = datetime.datetime.now().isoformat()

        # ワーカートークンで名前を補完
        token = fields.get("workerToken", "")
        worker_name = fields.get("name", "").strip()
        if not worker_name and token:
            w = get_worker_by_token_notion(token)
            if w: worker_name = w.get("name", "")

        invoice = {
            "id":             inv_id,
            "received_at":    now,
            "worker_name":    worker_name,
            "billing_month":  fields.get("billingMonth", ""),
            "amount":         int(fields.get("amount", 0) or 0),
            "status":         "new",
            "bank_name":      fields.get("bankName", ""),
            "branch":         fields.get("bankBranch", ""),
            "account_type":   fields.get("acctType", "普通"),
            "account_number": fields.get("bankNo", ""),
            "account_holder": fields.get("bankHolder", ""),
            "remark":         fields.get("remarks", ""),
            "has_pdf":        "pdf" in files,
            "worker_id":      "",
        }

        # PDF一時保存（Railwayでは再デプロイ時に消えるが、請求書データはNotionに永続保存）
        if "pdf" in files:
            _, pdf_bytes = files["pdf"]
            pdf_path = os.path.join(INVOICE_PDF_DIR, inv_id + ".pdf")
            try:
                with open(pdf_path, "wb") as f:
                    f.write(pdf_bytes)
            except Exception as e:
                print(f"  ⚠️ PDF一時保存失敗: {e}")

        # Notionに保存
        result, err = notion_request("POST", "/pages", {
            "parent": {"database_id": INVOICE_DB_ID},
            "properties": invoice_to_notion_props(invoice),
        })
        if not result:
            self.send_json(500, {"ok": False, "error": f"Notion保存失敗: {err}"}); return

        print(f"  ✅ 請求書Notion保存: {inv_id} ({worker_name} {invoice['billing_month']} ¥{invoice['amount']:,})")
        self.send_json(200, {"ok": True, "invoiceId": inv_id})

    # ── ワーカー認証（Notion）────────────────────────────────────────
    def handle_worker_login(self, data):
        worker_id = data.get("workerId", "").strip()
        password  = data.get("password", "")
        if not WORKER_DB_ID:
            self.send_json(500, {"ok": False, "error": "WORKER_DB_ID が設定されていません"}); return
        w = get_worker_by_id_notion(worker_id)
        if not w:
            self.send_json(401, {"ok": False, "error": "IDが見つかりません"}); return
        pw_hash = hashlib.sha256(password.encode()).hexdigest()
        if w.get("pw_hash") != pw_hash:
            self.send_json(401, {"ok": False, "error": "パスワードが違います"}); return
        # トークンがなければ新規発行してNotionに保存
        token = w.get("token") or str(uuid.uuid4())
        if not w.get("token"):
            notion_id = w.get("_notion_id")
            if notion_id:
                notion_request("PATCH", f"/pages/{notion_id}", {
                    "properties": {"トークン": {"rich_text": [{"text": {"content": token}}]}}
                })
        profile = {k: w.get(k) for k in ["name","bankName","bankBranch","bankNo","bankHolder","acctType"]}
        self.send_json(200, {"ok": True, "token": token, "profile": profile})

    def handle_worker_register(self, data):
        worker_id = data.get("workerId", "").strip()
        password  = data.get("password", "")
        name      = data.get("name", "").strip()
        if not worker_id or not password or not name:
            self.send_json(400, {"ok": False, "error": "必須項目が不足しています"}); return
        if len(password) < 8:
            self.send_json(400, {"ok": False, "error": "パスワードは8文字以上"}); return
        if not WORKER_DB_ID:
            self.send_json(500, {"ok": False, "error": "WORKER_DB_ID が設定されていません"}); return
        existing = get_worker_by_id_notion(worker_id)
        if existing:
            self.send_json(409, {"ok": False, "error": "このIDは既に使われています"}); return
        token   = str(uuid.uuid4())
        pw_hash = hashlib.sha256(password.encode()).hexdigest()
        new_w = {
            "worker_id": worker_id, "name": name, "pw_hash": pw_hash, "token": token,
            "bankName":"","bankBranch":"","bankNo":"","bankHolder":"","acctType":"普通","note":"",
        }
        result, err = notion_request("POST", "/pages", {
            "parent": {"database_id": WORKER_DB_ID},
            "properties": worker_to_notion_props(new_w),
        })
        if not result:
            self.send_json(500, {"ok": False, "error": f"Notion保存失敗: {err}"}); return
        profile = {k: new_w.get(k) for k in ["name","bankName","bankBranch","bankNo","bankHolder","acctType"]}
        self.send_json(200, {"ok": True, "token": token, "profile": profile})

    # ── ワーカー登録（請求書から）──────────────────────────
    def handle_register_worker_from_invoice(self, inv_id, data):
        worker_id = data.get("worker_id", "").strip()
        note      = data.get("note", "")
        if not worker_id:
            self.send_json(400, {"ok": False, "error": "ワーカーIDが必要"}); return

        # 請求書のNotionページIDを検索
        notion_page_id = find_invoice_notion_id(inv_id)
        if not notion_page_id:
            self.send_json(404, {"ok": False, "error": "請求書が見つかりません"}); return

        # 請求書のワーカーIDをNotionで更新
        notion_request("PATCH", f"/pages/{notion_page_id}", {
            "properties": {"ワーカーID": {"rich_text": [{"text": {"content": worker_id}}]}}
        })

        # 請求書データを取得してワーカー情報を補完
        inv_result, _ = notion_request("GET", f"/pages/{notion_page_id}")
        inv = parse_notion_invoice_page(inv_result) if inv_result else {}

        # ワーカーが未登録なら作成
        if WORKER_DB_ID:
            existing = get_worker_by_id_notion(worker_id)
            if not existing:
                new_w = {
                    "worker_id": worker_id,
                    "name":        inv.get("worker_name", ""),
                    "pw_hash":     "", "token": "",
                    "bankName":    inv.get("bank_name", ""),
                    "bankBranch":  inv.get("branch", ""),
                    "bankNo":      inv.get("account_number", ""),
                    "bankHolder":  inv.get("account_holder", ""),
                    "acctType":    inv.get("account_type", "普通"),
                    "note":        note,
                }
                notion_request("POST", "/pages", {
                    "parent": {"database_id": WORKER_DB_ID},
                    "properties": worker_to_notion_props(new_w),
                })
                print(f"  👤 ワーカー登録: {worker_id}")
        self.send_json(200, {"ok": True})

    def send_file(self, filepath, filename, content_type):
        from urllib.parse import quote
        with open(filepath, 'rb') as f:
            data = f.read()
        self.send_response(200)
        self.send_header('Content-Type', content_type)
        self.send_header('Content-Disposition', f"attachment; filename*=UTF-8''{quote(filename)}")
        self.send_header('Content-Length', str(len(data)))
        self.send_header('Access-Control-Allow-Origin', '*')
        self.send_header('Access-Control-Expose-Headers', 'Content-Disposition')
        self.end_headers()
        self.wfile.write(data)

    def handle_generate_invoice(self, data):
        if not OPENPYXL_OK:
            self.send_json(500, {'ok': False, 'error': 'openpyxl未インストール。pip install openpyxl --break-system-packages を実行してください'})
            return
        try:
            self._generate_invoice_inner(data)
        except Exception as e:
            import traceback
            traceback.print_exc()
            self.send_json(500, {'ok': False, 'error': f'請求書生成エラー: {e}'})

    def _generate_invoice_inner(self, data):
        customer_no   = data.get('customerNo', '')
        invoice_date  = data.get('invoiceDate', '')
        cases         = data.get('cases', [])

        # お客様名をNotionから取得（案件表のお客様no/名フィールド値で検索）
        customer_name = data.get('customerName', '')
        if not customer_name:
            body = {"filter": {"property": "お客様No.", "rich_text": {"equals": customer_no}}, "page_size": 1}
            res, _ = notion_request("POST", f"/databases/{CUSTOMER_DB_ID}/query", body)
            if res and res.get("results"):
                props = res["results"][0].get("properties", {})
                name_rt = props.get("クライアント名", {}).get("rich_text") or []
                customer_name = name_rt[0].get("plain_text", "") if name_rt else ""
            if not customer_name:
                customer_name = customer_no + "様"
        if not customer_name.endswith("様"):
            customer_name += "様"

        ORANGE = "B07C1A"
        BLUE   = "4472C4"

        def S(): return Side(border_style='thin', color='000000')
        def B(t=False, b=False, l=False, r=False):
            return Border(top=S() if t else Side(), bottom=S() if b else Side(),
                          left=S() if l else Side(), right=S() if r else Side())

        wb = Workbook()
        ws = wb.active
        ws.title = "請求書"
        ws.column_dimensions['A'].width = 1.5
        ws.column_dimensions['B'].width = 36
        ws.column_dimensions['C'].width = 14
        ws.column_dimensions['D'].width = 18

        def c(row, col, val=None, font=None, align=None, border=None, fmt=None):
            cell = ws.cell(row, col)
            if val is not None:  cell.value = val
            if font:   cell.font = font
            if align:  cell.alignment = align
            if border: cell.border = border
            if fmt:    cell.number_format = fmt
            return cell

        r = 1
        c(r,2,'請求書 No.', font=Font(name='メイリオ',size=8,color=BLUE)); r+=1
        ws.row_dimensions[r].height = 34
        c(r,2, customer_name, font=Font(name='メイリオ',size=22,color=ORANGE),
          align=Alignment(vertical='center'))
        ws.merge_cells(f'B{r}:C{r}'); r+=1
        ws.row_dimensions[r].height = 6; r+=1
        ws.row_dimensions[r].height = 6; r+=1
        c(r,2,'野津　欧', font=Font(name='メイリオ',size=11,color=ORANGE)); r+=1
        c(r,2,'三菱UFJ銀行新宿通り支店　（050）-0571808',
          font=Font(name='メイリオ',size=9,color=ORANGE)); r+=1
        ws.row_dimensions[r].height = 10; r+=1

        ws.row_dimensions[r].height = 20
        c(r,2, f'請求日：{invoice_date}',
          font=Font(name='メイリオ',size=10,color=ORANGE), border=B(t=True,l=True))
        c(r,3,'内容', font=Font(name='メイリオ',size=10,color=ORANGE,bold=True),
          align=Alignment(horizontal='center',vertical='center'), border=B(t=True,l=True,r=True)); r+=1

        ws.row_dimensions[r].height = 20
        c(r,2,'支払い期日：', font=Font(name='メイリオ',size=10), border=B(b=True,l=True))
        c(r,3,'動画編集案件の件', font=Font(name='メイリオ',size=10), border=B(b=True,l=True,r=True)); r+=1
        ws.row_dimensions[r].height = 8; r+=1

        ws.row_dimensions[r].height = 20
        c(r,2,'詳細', font=Font(name='メイリオ',size=10,color=ORANGE,bold=True), border=B(t=True,b=True,l=True))
        c(r,3,'金額', font=Font(name='メイリオ',size=10,color=ORANGE,bold=True),
          align=Alignment(horizontal='right',vertical='center'), border=B(t=True,b=True,l=True,r=True))
        c(r,4,'当方案件番号', font=Font(name='メイリオ',size=9,color=ORANGE)); r+=1

        total = 0
        for case in cases:
            desc    = case.get('note') or case.get('number','')
            amount  = int(case.get('amount') or case.get('price') or 0)
            case_no = case.get('number','')
            total  += amount
            ws.row_dimensions[r].height = 18
            c(r,2,desc, font=Font(name='メイリオ',size=10),
              align=Alignment(wrap_text=True,vertical='center'), border=B(b=True,l=True))
            c(r,3,amount, font=Font(name='メイリオ',size=10),
              align=Alignment(horizontal='right',vertical='center'),
              border=B(b=True,l=True,r=True), fmt='#,##0')
            c(r,4,case_no, font=Font(name='メイリオ',size=9,color='AAAAAA')); r+=1

        ws.row_dimensions[r].height = 8; r+=1

        for text, bold, bordered in [
            (f'小計　¥{total:,}',  False, False),
            ('税率　0%',           False, False),
            ('その他　¥0',         False, False),
            (f'集計　¥{total:,}',  True,  True),
        ]:
            ws.merge_cells(f'B{r}:C{r}')
            c(r,2,text, font=Font(name='メイリオ',size=10,color=ORANGE,bold=bold),
              align=Alignment(horizontal='center'),
              border=B(t=True,b=True,l=True,r=True) if bordered else None); r+=1

        ws.row_dimensions[r].height = 12; r+=1
        c(r,2,'この請求書に関してご不明な点がございましたら、お問い合わせください。',
          font=Font(name='メイリオ',size=9,color=ORANGE)); r+=2
        c(r,2,'今月もありがとうございます',
          font=Font(name='メイリオ',size=11,color=ORANGE,bold=True))

        ws.print_area = f'B1:C{r+1}'
        ws.page_margins.left = 0; ws.page_margins.right = 0
        ws.page_margins.top  = 0.4; ws.page_margins.bottom = 0.4

        tmp = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False)
        tmp.close()
        wb.save(tmp.name)
        safe = invoice_date.replace('/', '-')
        filename = f'{customer_no}_請求書_{safe}.xlsx'
        print(f"\n📄 請求書生成: {filename}")
        self.send_file(tmp.name, filename,
                       'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        os.unlink(tmp.name)

    # ========== カレンダー（案件表と完全同期）==========
    def handle_get_calendar(self):
        from urllib.parse import urlparse, parse_qs
        qs = parse_qs(urlparse(self.path).query)
        month_str = (qs.get('month') or [None])[0]
        today = datetime.date.today()
        if month_str:
            try:
                year  = int(month_str.split('-')[0])
                month = int(month_str.split('-')[1])
            except Exception:
                year, month = today.year, today.month
        else:
            year, month = today.year, today.month

        start = f"{year}-{month:02d}-01"
        end   = f"{year+1}-01-01" if month == 12 else f"{year}-{month+1:02d}-01"
        print(f"\n📅 カレンダー取得: {year}年{month}月")

        events = []
        # 案件表を全件取得（お客様あり=案件、なし=予定）
        q = {"filter": {"and": [
                {"property": "案件締切日・進行", "date": {"on_or_after": start}},
                {"property": "案件締切日・進行", "date": {"before": end}},
             ]}, "page_size": 100}
        cursor = None
        while True:
            if cursor:
                q["start_cursor"] = cursor
            result, _ = notion_request("POST", f"/databases/{CASE_DB_ID}/query", q)
            if not result:
                break
            for page in result.get("results", []):
                try:
                    props    = page["properties"]
                    number   = (props["当方案件番号"]["title"] or [{}])[0].get("plain_text","").strip()
                    customer = (props["お客様no/名"]["rich_text"] or [{}])[0].get("plain_text","").strip()
                    dl       = (props["案件締切日・進行"]["date"] or {}).get("start","")
                    status   = (props["進捗"]["status"] or {}).get("name","")
                    memo_rt  = (props.get("備考") or {}).get("rich_text") or []
                    memo     = memo_rt[0].get("plain_text","") if memo_rt else ""
                    if not dl:
                        continue
                    date_only = dl[:10]
                    # 時刻情報を抽出（例: "2026-03-30T07:30:00.000+09:00" → "07:30"）
                    start_time = dl[11:16] if len(dl) > 10 else ""
                    dl_end = (props["案件締切日・進行"]["date"] or {}).get("end","")
                    end_time = dl_end[11:16] if dl_end and len(dl_end) > 10 else ""
                    if customer:
                        # お客様ありは案件締切
                        events.append({"date": date_only, "title": number,
                                       "type": "deadline", "customer": customer,
                                       "status": status, "memo": memo, "id": page["id"]})
                    else:
                        # お客様なしは予定エントリー（バイト・会議等）
                        events.append({"date": date_only, "title": number,
                                       "type": "schedule", "memo": memo, "id": page["id"],
                                       "startTime": start_time, "endTime": end_time})
                except Exception:
                    pass
            if not result.get("has_more"):
                break
            cursor = result.get("next_cursor")

        self.send_json(200, {"ok": True,
                             "month": f"{year}-{month:02d}",
                             "events": events})

    def handle_get_calendar_day(self):
        """特定日の予定・案件をすべて返す（日次スケジュール用）"""
        from urllib.parse import urlparse, parse_qs
        qs   = parse_qs(urlparse(self.path).query)
        date = (qs.get("date") or [None])[0]
        if not date:
            self.send_json(400, {"ok": False, "error": "dateが必要"}); return
        print(f"\n📅 カレンダー日別取得: {date}")
        next_day = str(datetime.date.fromisoformat(date) + datetime.timedelta(days=1))
        q = {"filter": {"and": [
                {"property": "案件締切日・進行", "date": {"on_or_after": date}},
                {"property": "案件締切日・進行", "date": {"before": next_day}},
             ]}, "page_size": 100}
        result, _ = notion_request("POST", f"/databases/{CASE_DB_ID}/query", q)
        events = []
        for page in (result or {}).get("results", []):
            try:
                props    = page["properties"]
                number   = (props["当方案件番号"]["title"] or [{}])[0].get("plain_text","").strip()
                customer = (props["お客様no/名"]["rich_text"] or [{}])[0].get("plain_text","").strip()
                dl       = (props["案件締切日・進行"]["date"] or {}).get("start","")
                dl_end   = (props["案件締切日・進行"]["date"] or {}).get("end","")
                status   = (props["進捗"]["status"] or {}).get("name","")
                memo_rt  = (props.get("備考") or {}).get("rich_text") or []
                memo     = memo_rt[0].get("plain_text","") if memo_rt else ""
                if not dl: continue
                start_time = dl[11:16] if len(dl) > 10 else ""
                end_time   = dl_end[11:16] if dl_end and len(dl_end) > 10 else ""
                if customer:
                    events.append({"date": dl[:10], "title": number,
                                   "type": "deadline", "customer": customer,
                                   "status": status, "memo": memo, "id": page["id"]})
                else:
                    events.append({"date": dl[:10], "title": number,
                                   "type": "schedule", "memo": memo, "id": page["id"],
                                   "startTime": start_time, "endTime": end_time})
            except Exception:
                pass
        self.send_json(200, {"ok": True, "date": date, "events": events})

    # ========== 引き継ぎ CRUD（Notion DB）==========
    def handle_get_handover_all(self):
        """全アクティブ引き継ぎを日付順で返す"""
        body = {"filter": {"property": "ステータス", "select": {"equals": "active"}},
                "page_size": 100}
        result, err = notion_request("POST", f"/databases/{HANDOVER_DB_ID}/query", body)
        if not result:
            self.send_json(500, {"ok": False, "error": err}); return
        items = [i for i in (parse_handover_page(p) for p in result.get("results",[])) if i]
        items.sort(key=lambda i: i.get("start_date", i.get("date","9999-12-31")))
        self.send_json(200, {"ok": True, "items": items})

    def handle_handover_delete(self, data):
        """引き継ぎをアーカイブ（削除）"""
        item_id = data.get("id","")
        notion_request("PATCH", f"/pages/{item_id}", {"archived": True})
        print(f"  🗑️  引き継ぎ削除: {item_id[:8]}")
        self.send_json(200, {"ok": True})

    def handle_get_handover(self):
        from urllib.parse import urlparse, parse_qs
        qs = parse_qs(urlparse(self.path).query)
        date_str = (qs.get('date') or [None])[0] or datetime.date.today().isoformat()
        items = get_active_handover(date_str)
        self.send_json(200, {"ok": True, "items": items, "date": date_str})

    def handle_handover_add(self, data):
        text = data.get("text","").strip()
        typ  = data.get("type","todo")
        if not text:
            self.send_json(400, {"ok": False, "error": "テキストが必要"}); return
        today    = datetime.date.today().isoformat()
        tomorrow = (datetime.date.today() + datetime.timedelta(days=1)).isoformat()
        props = {
            "内容":     {"title":  [{"text": {"content": text}}]},
            "種別":     {"select": {"name": typ}},
            "ステータス": {"select": {"name": "active"}},
            "作成日":   {"date":   {"start": today}},
        }
        if typ == "medium":
            props["開始日"] = {"date": {"start": data.get("start_date", today)}}
            props["終了日"] = {"date": {"start": data.get("end_date",   today)}}
        else:
            props["日付"] = {"date": {"start": data.get("date", tomorrow)}}
        result, err = notion_request("POST", "/pages",
                                     {"parent": {"database_id": HANDOVER_DB_ID},
                                      "properties": props})
        if not result:
            self.send_json(400, {"ok": False, "error": err}); return
        item = parse_handover_page(result)
        print(f"  ✅ 引き継ぎ追加: [{typ}] {text[:30]}")
        self.send_json(200, {"ok": True, "item": item})

    def handle_handover_done(self, data):
        item_id = data.get("id","")
        notion_request("PATCH", f"/pages/{item_id}",
                       {"properties": {"ステータス": {"select": {"name": "done"}}}})
        print(f"  ✅ 引き継ぎ完了: {item_id[:8]}")
        self.send_json(200, {"ok": True})

    def handle_handover_carry(self, data):
        """翌日へ引き継ぎ（種別を変えて再登録）"""
        item_id = data.get("id","")
        as_type = data.get("as_type","todo")
        to_date = data.get("to_date",
                           (datetime.date.today() + datetime.timedelta(days=1)).isoformat())
        # 元アイテムを取得して完了にする
        src_result, _ = notion_request("GET", f"/pages/{item_id}")
        notion_request("PATCH", f"/pages/{item_id}",
                       {"properties": {"ステータス": {"select": {"name": "done"}}}})
        if src_result:
            src = parse_handover_page(src_result)
            if src:
                props = {
                    "内容":     {"title":  [{"text": {"content": src["text"]}}]},
                    "種別":     {"select": {"name": as_type}},
                    "ステータス": {"select": {"name": "active"}},
                    "作成日":   {"date":   {"start": datetime.date.today().isoformat()}},
                    "日付":     {"date":   {"start": to_date}},
                }
                notion_request("POST", "/pages",
                               {"parent": {"database_id": HANDOVER_DB_ID},
                                "properties": props})
                print(f"  🔁 引き継ぎ → {to_date} [{as_type}] {src['text'][:30]}")
        self.send_json(200, {"ok": True})

    def handle_handover_extend(self, data):
        """中期メモの終了日を延長"""
        item_id = data.get("id","")
        new_end  = data.get("end_date","")
        notion_request("PATCH", f"/pages/{item_id}",
                       {"properties": {"終了日": {"date": {"start": new_end}}}})
        print(f"  📅 中期延長: {item_id[:8]} → {new_end}")
        self.send_json(200, {"ok": True})

    def handle_handover_update_date(self, data):
        """日付変更"""
        item_id = data.get("id","")
        props = {}
        if data.get("start_date"): props["開始日"] = {"date": {"start": data["start_date"]}}
        if data.get("end_date"):   props["終了日"] = {"date": {"start": data["end_date"]}}
        if data.get("date"):       props["日付"]   = {"date": {"start": data["date"]}}
        if props:
            notion_request("PATCH", f"/pages/{item_id}", {"properties": props})
        print(f"  📅 日付変更: {item_id[:8]}")
        self.send_json(200, {"ok": True})

    def handle_get_handover_done_list(self):
        """完了済み引き継ぎ一覧を返す"""
        items = get_done_handover()
        self.send_json(200, {"ok": True, "items": items})

    def handle_handover_restore(self, data):
        """完了済み引き継ぎをアクティブに戻す"""
        item_id = data.get("id", "")
        if not item_id:
            self.send_json(400, {"ok": False, "error": "idが必要"}); return
        notion_request("PATCH", f"/pages/{item_id}",
                       {"properties": {"ステータス": {"select": {"name": "active"}}}})
        print(f"  🔄 引き継ぎ復元: {item_id[:8]}")
        self.send_json(200, {"ok": True})

    def handle_handover_update_content(self, data):
        """引き継ぎ内容テキストを更新"""
        item_id = data.get("id", "")
        text    = data.get("text", "").strip()
        if not item_id or not text:
            self.send_json(400, {"ok": False, "error": "idとtextが必要"}); return
        notion_request("PATCH", f"/pages/{item_id}", {
            "properties": {"内容": {"title": [{"text": {"content": text}}]}}
        })
        print(f"  ✏️  内容更新: {item_id[:8]} → {text[:20]}")
        self.send_json(200, {"ok": True})

    def handle_get_schedule_dates(self):
        """指定月に日次スケジュールが保存されている日付一覧を返す"""
        from urllib.parse import urlparse, parse_qs
        qs    = parse_qs(urlparse(self.path).query)
        month = (qs.get("month") or [None])[0]  # "YYYY-MM"
        if not month:
            self.send_json(400, {"ok": False, "error": "monthが必要"}); return
        schedules = load_daily_schedules()
        # ブロックが1件以上ある日のみ返す
        dates = [d for d in schedules if d.startswith(month) and len(schedules[d]) > 0]
        self.send_json(200, {"ok": True, "dates": dates})

    def handle_get_daily_schedule(self):
        """日次スケジュールブロックを返す"""
        from urllib.parse import urlparse, parse_qs
        qs   = parse_qs(urlparse(self.path).query)
        date = qs.get("date", [None])[0]
        if not date:
            self.send_json(400, {"ok": False, "error": "dateが必要"}); return
        schedules = load_daily_schedules()
        blocks = schedules.get(date, [])
        self.send_json(200, {"ok": True, "date": date, "blocks": blocks})

    def handle_save_daily_block(self, data):
        """日次スケジュールブロックを保存/更新"""
        import uuid
        date  = data.get("date", "")
        block = data.get("block", {})
        if not date or not block:
            self.send_json(400, {"ok": False, "error": "dateとblockが必要"}); return
        if not block.get("id"):
            block["id"] = str(uuid.uuid4())[:8]
        schedules = load_daily_schedules()
        blocks = schedules.get(date, [])
        # 同IDがあれば更新、なければ追加
        updated = False
        for i, b in enumerate(blocks):
            if b.get("id") == block["id"]:
                blocks[i] = block; updated = True; break
        if not updated:
            blocks.append(block)
        schedules[date] = blocks
        save_daily_schedules(schedules)
        print(f"  📅 日次ブロック保存: {date} id={block['id']}")
        self.send_json(200, {"ok": True, "block": block})

    def handle_delete_daily_block(self, data):
        """日次スケジュールブロックを削除"""
        date     = data.get("date", "")
        block_id = data.get("id", "")
        if not date or not block_id:
            self.send_json(400, {"ok": False, "error": "dateとidが必要"}); return
        schedules = load_daily_schedules()
        blocks = schedules.get(date, [])
        schedules[date] = [b for b in blocks if b.get("id") != block_id]
        save_daily_schedules(schedules)
        print(f"  🗑️  日次ブロック削除: {date} id={block_id}")
        self.send_json(200, {"ok": True})

    def handle_move_notion_event(self, data):
        """Notion予定ブロックの時刻を更新"""
        page_id   = data.get("id", "")
        date      = data.get("date", "")
        start_min = int(data.get("startMin", 0))
        end_min   = int(data.get("endMin", 60))
        if not page_id or not date:
            self.send_json(400, {"ok": False, "error": "id/dateが必要"}); return
        sh = f"{start_min // 60:02d}:{start_min % 60:02d}"
        eh = f"{end_min   // 60:02d}:{end_min   % 60:02d}"
        start_dt = f"{date}T{sh}:00+09:00"
        end_dt   = f"{date}T{eh}:00+09:00"
        props = {"案件締切日・進行": {"date": {"start": start_dt, "end": end_dt}}}
        result, err = notion_request("PATCH", f"/pages/{page_id}", {"properties": props})
        if err:
            self.send_json(500, {"ok": False, "error": err}); return
        print(f"  📅 Notion予定移動: {page_id[:8]} {sh}〜{eh}")
        self.send_json(200, {"ok": True, "startTime": sh, "endTime": eh})

    def handle_get_cases_list(self):
        """全案件を検索・フィルタして返す（案件履歴ツール用）"""
        from urllib.parse import urlparse, parse_qs
        qs = parse_qs(urlparse(self.path).query)
        q_term   = (qs.get("q")        or [""])[0].strip()
        status_f = (qs.get("status")   or [""])[0].strip()
        cust_f   = (qs.get("customer") or [""])[0].strip()

        month_f  = (qs.get("month")    or [""])[0].strip()   # e.g. "2026-03"

        filters = []
        # 月フィルター（デフォルト: 当月）
        if month_f:
            try:
                y, m = map(int, month_f.split('-'))
                ms = f"{y}-{m:02d}-01"
                me = f"{y+1}-01-01" if m == 12 else f"{y}-{m+1:02d}-01"
                filters.append({"property": "案件締切日・進行", "date": {"on_or_after": ms}})
                filters.append({"property": "案件締切日・進行", "date": {"before": me}})
            except Exception:
                pass
        if cust_f:
            filters.append({"property": "お客様no/名", "rich_text": {"contains": cust_f}})
        if status_f:
            filters.append({"property": "進捗", "status": {"equals": status_f}})
        # キーワード検索は案件番号のみ（Notion APIのorフィルター制限を回避）
        if q_term:
            filters.append({"property": "当方案件番号", "title": {"contains": q_term}})

        query_body = {
            "page_size": 100,
            "sorts": [{"property": "案件締切日・進行", "direction": "descending"}],
        }
        if len(filters) == 1:   query_body["filter"] = filters[0]
        elif len(filters) > 1:  query_body["filter"] = {"and": filters}

        cases = []
        cursor = None
        while True:
            if cursor: query_body["start_cursor"] = cursor
            result, err = notion_request("POST", f"/databases/{CASE_DB_ID}/query", query_body)
            if not result:
                self.send_json(400, {"ok": False, "error": err}); return
            for page in result.get("results", []):
                try:
                    props    = page["properties"]
                    number   = (props["当方案件番号"]["title"]      or [{}])[0].get("plain_text","").strip()
                    customer = (props["お客様no/名"]["rich_text"]   or [{}])[0].get("plain_text","").strip()
                    price_v  = props["単価"]["number"]
                    price    = price_v if price_v is not None else 0
                    note     = (props.get("備考/素材名",{}).get("rich_text") or [{}])[0].get("plain_text","")
                    filename = (props.get("指定案件ファイル名",{}).get("rich_text") or [{}])[0].get("plain_text","")
                    memo     = (props.get("備考",{}).get("rich_text")        or [{}])[0].get("plain_text","")
                    gross    = (props.get("粗利（単価-外注費）",{}).get("rich_text") or [{}])[0].get("plain_text","")
                    dl       = (props.get("案件締切日・進行",{}).get("date") or {}).get("start","")
                    status   = (props.get("進捗",{}).get("status") or {}).get("name","")
                    # 粗利テキストから利益額を抽出（例: "10,000 - 3,000 = 7,000" → 7000）
                    # grossが空の場合は外注費ゼロとみなし粗利=単価として扱う
                    import re as _re
                    profit = None
                    if gross:
                        m = _re.search(r'=\s*([\d,]+)', gross)
                        if m:
                            try: profit = int(m.group(1).replace(',',''))
                            except: pass
                    if profit is None and price:
                        profit = int(price)   # 外注費未入力 → 粗利=単価
                    if not customer:
                        continue   # 客番号なし（バイト・会議等の予定）は除外
                    cases.append({
                        "id": page["id"], "number": number, "customer": customer,
                        "price": price, "note": note, "filename": filename,
                        "memo": memo, "gross": gross, "profit": profit,
                        "date": dl[:10] if dl else "", "status": status,
                        "url": page.get("url",""),
                    })
                except Exception:
                    pass
            if not result.get("has_more"): break
            cursor = result.get("next_cursor")

        # 顧客名を付与（キャッシュから引く）
        name_map = get_customer_name_map()
        for c in cases:
            c["customerName"] = name_map.get(c["customer"], "")

        print(f"\n📋 案件履歴取得: {len(cases)}件 (q={q_term!r} status={status_f!r} cust={cust_f!r})")
        self.send_json(200, {"ok": True, "cases": cases, "total": len(cases)})

    def handle_update_case(self, data):
        """案件のプロパティを更新"""
        page_id = data.get("id","")
        if not page_id:
            self.send_json(400, {"ok": False, "error": "idが必要"}); return
        props = {}
        if "number"   in data: props["当方案件番号"]        = {"title": [{"text": {"content": data["number"]}}]}
        if "customer" in data: props["お客様no/名"]         = {"rich_text": [{"text": {"content": data["customer"]}}]}
        if "note"     in data: props["備考/素材名"]          = {"rich_text": [{"text": {"content": data["note"]}}]}
        if "memo"     in data: props["備考"]                = {"rich_text": [{"text": {"content": data["memo"]}}]}
        if "filename" in data: props["指定案件ファイル名"]  = {"rich_text": [{"text": {"content": data["filename"]}}]}
        if "status"   in data: props["進捗"]                = {"status": {"name": data["status"]}}
        if "date"     in data:
            props["案件締切日・進行"] = {"date": {"start": data["date"]}} if data["date"] else {"date": None}
        if "price" in data:
            try:
                pv = float(data["price"]) if str(data["price"]).strip() != "" else 0.0
                props["単価"] = {"number": pv}
                # 外注費：空・未送信はすべて0扱い
                raw_cost = str(data.get("outsourceCost", "")).strip()
                cv = float(raw_cost) if raw_cost != "" else 0.0
                gross_text = f"{int(pv):,} - {int(cv):,} = {int(pv - cv):,}"
                props["粗利（単価-外注費）"] = {"rich_text": [{"text": {"content": gross_text}}]}
            except Exception as e:
                print(f"  ⚠️ 粗利計算エラー: {e}")
        if not props:
            self.send_json(400, {"ok": False, "error": "更新項目なし"}); return
        result, err = notion_request("PATCH", f"/pages/{page_id}", {"properties": props})
        if result:
            print(f"  ✅ 案件更新: {page_id[:8]} → {list(props.keys())}")
            self.send_json(200, {"ok": True})
        else:
            self.send_json(500, {"ok": False, "error": err})

    def handle_gantt_excel(self, data):
        """ガントチャートExcelを生成して返す"""
        import tempfile
        from openpyxl import Workbook
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        date_str = data.get("date", "")
        blocks   = data.get("blocks", [])
        notion   = data.get("notionEvents", [])

        def time_to_min(t):
            try:
                h, m = t.split(":")
                return int(h) * 60 + int(m)
            except Exception:
                return 0

        SLOT_MINS = 30
        SLOTS     = 24 * 60 // SLOT_MINS  # 48列

        TYPE_COLORS = {
            "custom":   "7C3AED",
            "todo":     "2563EB",
            "memo":     "D97706",
            "deadline": "DC2626",
            "medium":   "059669",
            "notion":   "374151",
        }
        TYPE_LABELS = {
            "custom": "作業", "todo": "TODO", "memo": "メモ",
            "deadline": "納期", "medium": "中期", "notion": "Notion予定",
        }

        entries = []
        for b in blocks:
            entries.append({
                "text":  b.get("text") or "(無題)",
                "type":  b.get("type", "custom"),
                "start": b.get("startMin", 0),
                "end":   b.get("endMin",   b.get("startMin", 0) + 60),
                "notion_linked": bool(b.get("notionId")),
            })
        for ev in notion:
            st = time_to_min(ev.get("startTime", ""))
            en = time_to_min(ev.get("endTime",   "")) if ev.get("endTime") else st + 60
            entries.append({
                "text":  ev.get("title") or "(無題)",
                "type":  "notion",
                "start": st,
                "end":   en,
                "notion_linked": True,
            })
        entries.sort(key=lambda e: e["start"])

        wb = Workbook()
        ws = wb.active
        ws.title = "ガントチャート"

        HEADER_FILL = PatternFill("solid", start_color="1E1B4B", end_color="1E1B4B")
        HOUR_FILL   = PatternFill("solid", start_color="312E81", end_color="312E81")
        EVEN_ROW    = PatternFill("solid", start_color="F8F9FF", end_color="F8F9FF")
        ODD_ROW     = PatternFill("solid", start_color="FFFFFF", end_color="FFFFFF")
        thin        = Side(border_style="thin", color="D1D5DB")
        bdr         = Border(left=thin, right=thin, top=thin, bottom=thin)

        def af(sz=9, bold=False, color="FFFFFF"):
            return Font(name="Arial", size=sz, bold=bold, color=color)

        # 列オフセット：A=内容, B=種別, C=時刻, D以降=ガントバー
        BAR_OFFSET = 4   # ガントバーが始まる列番号（1-indexed）

        # 1行目：タイトル（A〜最終ガント列まで結合）
        ws.merge_cells(f"A1:{get_column_letter(SLOTS + BAR_OFFSET - 1)}1")
        tc = ws["A1"]
        tc.value     = f"日次スケジュール　ガントチャート　{date_str}"
        tc.font      = af(11, True)
        tc.fill      = HEADER_FILL
        tc.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 22

        # 2行目：固定ヘッダー（内容／種別／時刻）
        for col, label in [(1, "内容"), (2, "種別"), (3, "時刻")]:
            c = ws.cell(2, col)
            c.value     = label
            c.font      = af(9, True)
            c.fill      = HEADER_FILL
            c.alignment = Alignment(horizontal="center", vertical="center")

        # 2行目：24時間軸ラベル（1時間ごとに表示）
        for i in range(SLOTS):
            mins = i * SLOT_MINS
            h, m = divmod(mins, 60)
            c = ws.cell(2, i + BAR_OFFSET)
            c.value     = f"{h:02d}:00" if m == 0 else ""
            c.font      = af(8, bold=(m == 0))
            c.fill      = HOUR_FILL if m == 0 else HEADER_FILL
            c.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[2].height = 16

        # 列幅
        ws.column_dimensions["A"].width = 22
        ws.column_dimensions["B"].width = 9
        ws.column_dimensions["C"].width = 13   # 時刻列
        for i in range(SLOTS):
            ws.column_dimensions[get_column_letter(i + BAR_OFFSET)].width = 2.2

        # データ行
        for ri, entry in enumerate(entries):
            row   = ri + 3
            bfill = EVEN_ROW if ri % 2 == 0 else ODD_ROW
            color = TYPE_COLORS.get(entry["type"], "7C3AED")
            gfill = PatternFill("solid", start_color=color, end_color=color)

            # 内容
            nc = ws.cell(row, 1)
            nc.value     = ("📅 " if entry["notion_linked"] else "") + entry["text"]
            nc.font      = Font(name="Arial", size=9)
            nc.fill      = bfill
            nc.alignment = Alignment(vertical="center")
            nc.border    = bdr

            # 種別
            sc = ws.cell(row, 2)
            sc.value     = TYPE_LABELS.get(entry["type"], entry["type"])
            sc.font      = Font(name="Arial", size=8)
            sc.fill      = bfill
            sc.alignment = Alignment(horizontal="center", vertical="center")
            sc.border    = bdr

            # 時刻（専用列）
            sh = f"{entry['start'] // 60:02d}:{entry['start'] % 60:02d}"
            eh = f"{entry['end']   // 60:02d}:{entry['end']   % 60:02d}"
            tc = ws.cell(row, 3)
            tc.value     = f"{sh}〜{eh}"
            tc.font      = Font(name="Arial", size=9, bold=True, color="111827")
            tc.fill      = bfill
            tc.alignment = Alignment(horizontal="center", vertical="center")
            tc.border    = bdr

            # ガントバー
            start_slot = entry["start"] // SLOT_MINS
            end_slot   = (entry["end"] + SLOT_MINS - 1) // SLOT_MINS
            for s in range(SLOTS):
                c = ws.cell(row, s + BAR_OFFSET)
                c.fill   = gfill if start_slot <= s < end_slot else bfill
                c.border = bdr
            ws.row_dimensions[row].height = 16

        if not entries:
            ws.cell(3, 1).value = "予定なし"
            ws.cell(3, 1).font  = Font(name="Arial", size=9, color="9CA3AF")

        tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
        tmp.close()
        wb.save(tmp.name)

        safe_date = date_str.replace("-", "")
        filename  = f"ガントチャート_{safe_date}.xlsx"
        print(f"\n📊 ガントチャート生成: {filename}  ({len(entries)}件)")
        self.send_file(tmp.name, filename,
                       "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        os.unlink(tmp.name)

    def handle_invoice_numbers(self, data):
        """Numbers.appにAppleScriptファイル経由でテンプレートを書き換えて請求書を生成"""
        import shutil, calendar, re, subprocess, tempfile

        if not shutil.which("osascript"):
            self.send_json(500, {"ok": False, "error": "osascriptが見つかりません。macOS上でサーバーを起動してください"}); return

        template_path = os.path.join(SCRIPT_DIR, "請求書雛形のコピー.numbers")
        if not os.path.exists(template_path):
            self.send_json(404, {"ok": False, "error": "請求書雛形のコピー.numbersが見つかりません"}); return

        customer_name = data.get("customerName", "")
        invoice_date  = data.get("invoiceDate", "")
        cases         = data.get("cases", [])
        customer_no   = data.get("customerNo", "")

        # 支払い期日（翌月末）
        due_date = ""
        md = re.match(r"(\d+)年(\d+)月(\d+)日", invoice_date)
        if md:
            y, mo = int(md.group(1)), int(md.group(2))
            nm = mo + 1 if mo < 12 else 1
            ny = y if mo < 12 else y + 1
            due_date = f"{ny}年{nm}月{calendar.monthrange(ny, nm)[1]}日"

        content_parts = [c.get("note") or c.get("number", "") for c in cases
                         if c.get("note") or c.get("number")]
        content_str = "、".join(content_parts) or "動画編集案件の件"

        total = sum(int(float(c.get("amount") or c.get("price") or 0)) for c in cases)

        def qs(s):
            """AppleScript文字列用クォートエスケープ"""
            return str(s).replace("\\", "\\\\").replace('"', '\\"')

        # テンプレート構造: データ行13〜14、小計15行目
        # 3件以上の場合は小計行(15)の直前に1行ずつ挿入
        TEMPLATE_ROWS = 2
        extra_rows   = max(0, len(cases) - TEMPLATE_ROWS)
        subtotal_row = 15 + extra_rows

        # 行挿入ブロック（挿入ごとに delay 0.5 で Numbers が処理を完了するのを待つ）
        insert_lines = []
        for i in range(extra_rows):
            insert_lines.append(f'                    add row above row {15 + i}')
            insert_lines.append(f'                    delay 0.5')
        insert_block = "\n".join(insert_lines) + "\n" if insert_lines else ""

        # データ行セット
        row_lines = []
        for i, c in enumerate(cases):
            row    = 13 + i
            detail = qs(c.get("note") or c.get("number", ""))
            amount = int(float(c.get("amount") or c.get("price") or 0))
            row_lines.append(f'                    set value of cell "B{row}" to "{detail}"')
            row_lines.append(f'                    set value of cell "C{row}" to {amount}')
        row_block = "\n".join(row_lines)

        # 一時ディレクトリ（ASCIIパス）
        tmp_dir   = tempfile.mkdtemp(prefix="inv_")
        tmp_path  = os.path.join(tmp_dir, "invoice.numbers")
        scpt_path = os.path.join(tmp_dir, "fill.applescript")
        shutil.copy(template_path, tmp_path)

        # AppleScriptをUTF-8ファイルとして書き出し（-e フラグだと日本語が文字化けするため）
        scpt = f'''\
-- Numbers請求書自動記入スクリプト
with timeout of 120 seconds
    set docFile to POSIX file "{tmp_path}"
    tell application "Numbers"
        set wasRunning to running
        set theDoc to open docFile
        -- ファイルが完全に開くまで待機
        delay 5
        tell theDoc
            tell sheet 1
                tell table 1
{insert_block}                    set value of cell "B2" to "{qs(customer_name)}様"
                    set value of cell "B8" to "請求日：{qs(invoice_date)}"
                    set value of cell "B9" to "支払い期日：{qs(due_date)}"
                    set value of cell "C9" to "{qs(content_str)}"
{row_block}
                    set value of cell "C{subtotal_row}" to {total}
                end tell
            end tell
            save
            delay 1
        end tell
        close theDoc saving no
        if not wasRunning then quit
    end tell
end timeout
'''
        with open(scpt_path, "w", encoding="utf-8") as f:
            f.write(scpt)

        print(f"\n📄 Numbers AppleScript:\n{scpt}")

        try:
            result = subprocess.run(
                ["osascript", scpt_path],
                capture_output=True, text=True, timeout=150
            )
            if result.stdout.strip():
                print(f"  stdout: {result.stdout.strip()}")
            if result.stderr.strip():
                print(f"  stderr: {result.stderr.strip()}")
            if result.returncode != 0:
                raise Exception(f"AppleScript失敗: {result.stderr.strip()[:600]}")

            safe_date = invoice_date.replace("年","").replace("月","").replace("日","")
            filename  = f"{customer_no}_請求書_{safe_date}.numbers"
            print(f"  ✅ {filename}")
            self.send_file(tmp_path, filename, "application/x-iwork-numbers-sffnumbers")
        except subprocess.TimeoutExpired:
            print("  ❌ タイムアウト（Numbers起動に150秒以上かかりました）")
            self.send_json(500, {"ok": False, "error": "タイムアウト：Numbers.appの起動に時間がかかりすぎました"})
        except Exception as e:
            print(f"  ❌ Numbers生成エラー: {e}")
            self.send_json(500, {"ok": False, "error": str(e)})
        finally:
            shutil.rmtree(tmp_dir, ignore_errors=True)

    def handle_invoice_template(self, data):
        """Numbers雛形と同じレイアウトでXLSX請求書をゼロから生成（外部テンプレート不要）"""
        import calendar, re
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

        customer_name = data.get("customerName", "")
        invoice_date  = data.get("invoiceDate", "")
        cases         = data.get("cases", [])
        customer_no   = data.get("customerNo", "")

        # 支払い期日：次月末
        due_date = ""
        m = re.match(r"(\d+)年(\d+)月(\d+)日", invoice_date)
        if m:
            y, mo = int(m.group(1)), int(m.group(2))
            nm = mo + 1 if mo < 12 else 1
            ny = y if mo < 12 else y + 1
            due_date = f"{ny}年{nm}月{calendar.monthrange(ny, nm)[1]}日"

        content_parts = [c.get("note") or c.get("number", "") for c in cases
                         if c.get("note") or c.get("number")]
        content_str = "、".join(content_parts) or "動画編集案件の件"
        total = sum(int(float(c.get("amount") or c.get("price") or 0)) for c in cases)

        # ---- ワークブック作成 ----
        wb = Workbook()
        ws = wb.active
        ws.title = "請求書"

        # ---- ヘルパー ----
        def S(style='thin'):
            return Side(style=style, color='000000')
        def B(t=None, r=None, b=None, l=None):
            return Border(
                top    = S(t) if t else Side(style=None),
                right  = S(r) if r else Side(style=None),
                bottom = S(b) if b else Side(style=None),
                left   = S(l) if l else Side(style=None),
            )
        def fill(hex_color):
            return PatternFill(fill_type='solid', fgColor=hex_color)

        GRAY  = 'D9D9D9'
        GOLD  = 'FFF2CC'

        # ---- 列幅（Numbers雛形に合わせた比率） ----
        ws.column_dimensions['A'].width = 2
        ws.column_dimensions['B'].width = 32
        ws.column_dimensions['C'].width = 18
        ws.column_dimensions['D'].width = 2

        # ---- Row 1: タイトル ----
        ws.row_dimensions[1].height = 36
        ws.merge_cells('B1:C1')
        c = ws['B1']
        c.value     = '請　求　書'
        c.font      = Font(size=22, bold=True)
        c.alignment = Alignment(horizontal='center', vertical='center')

        # ---- Row 2: 宛名 ----
        ws.row_dimensions[2].height = 28
        c = ws['B2']
        c.value     = f'{customer_name}　様'
        c.font      = Font(size=14, bold=True)
        c.alignment = Alignment(vertical='center')
        c.border    = B(b='medium')

        # ---- Rows 3-7: 空白（会社名等スペース） ----
        for r in range(3, 8):
            ws.row_dimensions[r].height = 14

        # ---- Row 8: 請求日 ＋ 「内容」ラベル ----
        ws.row_dimensions[8].height = 18
        c = ws['B8']
        c.value     = f'請求日：{invoice_date}'
        c.font      = Font(size=11)
        c.alignment = Alignment(vertical='center')
        c = ws['C8']
        c.value     = '内　容'
        c.font      = Font(size=10, bold=True)
        c.alignment = Alignment(horizontal='center', vertical='center')
        c.border    = B(t='thin', r='thin', b='thin', l='thin')
        c.fill      = fill(GRAY)

        # ---- Row 9: 支払期日 ＋ 内容テキスト ----
        ws.row_dimensions[9].height = 18
        c = ws['B9']
        c.value     = f'支払い期日：{due_date}'
        c.font      = Font(size=11)
        c.alignment = Alignment(vertical='center')
        c = ws['C9']
        c.value     = content_str
        c.font      = Font(size=10)
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        c.border    = B(r='thin', b='thin', l='thin')

        # ---- Rows 10-11: 空白 ----
        for r in [10, 11]:
            ws.row_dimensions[r].height = 10

        # ---- Row 12: 明細ヘッダー ----
        ws.row_dimensions[12].height = 20
        for col, label in [('B', '項　目'), ('C', '金　額（円）')]:
            c = ws[f'{col}12']
            c.value     = label
            c.font      = Font(bold=True, size=10)
            c.alignment = Alignment(horizontal='center', vertical='center')
            c.border    = B(t='medium', r='thin', b='thin', l='thin')
            c.fill      = fill(GRAY)

        # ---- Rows 13+: 明細データ ----
        for i, case in enumerate(cases):
            r      = 13 + i
            note   = case.get("note") or case.get("number", "")
            amount = int(float(case.get("amount") or case.get("price") or 0))
            ws.row_dimensions[r].height = 18
            bc = ws[f'B{r}']
            bc.value     = note
            bc.font      = Font(size=10)
            bc.alignment = Alignment(vertical='center')
            bc.border    = B(t='thin', r='thin', b='thin', l='thin')
            cc = ws[f'C{r}']
            cc.value         = amount
            cc.font          = Font(size=10)
            cc.number_format = '#,##0'
            cc.alignment     = Alignment(horizontal='right', vertical='center')
            cc.border        = B(t='thin', r='thin', b='thin', l='thin')

        # ---- 小計行 ----
        sr = 13 + len(cases)
        ws.row_dimensions[sr].height = 18
        c = ws[f'B{sr}']
        c.value     = '小　計'
        c.font      = Font(bold=True, size=10)
        c.alignment = Alignment(vertical='center')
        c.border    = B(t='thin', r='thin', b='thin', l='thin')
        c = ws[f'C{sr}']
        c.value         = total
        c.font          = Font(bold=True, size=10)
        c.number_format = '#,##0'
        c.alignment     = Alignment(horizontal='right', vertical='center')
        c.border        = B(t='thin', r='thin', b='thin', l='thin')

        # ---- 消費税・調整欄（Numbers の C16/C17 相当：空白） ----
        for dr in [sr + 1, sr + 2]:
            ws.row_dimensions[dr].height = 18
            ws[f'B{dr}'].border = B(r='thin', b='thin', l='thin')
            ws[f'C{dr}'].border = B(r='thin', b='thin', l='thin')

        # ---- 合計行（Numbers の C18 相当） ----
        tr = sr + 3
        ws.row_dimensions[tr].height = 24
        c = ws[f'B{tr}']
        c.value     = '合　計'
        c.font      = Font(bold=True, size=13)
        c.alignment = Alignment(vertical='center')
        c.border    = B(t='medium', r='thin', b='medium', l='medium')
        c.fill      = fill(GOLD)
        c = ws[f'C{tr}']
        c.value         = total
        c.font          = Font(bold=True, size=13)
        c.number_format = '#,##0'
        c.alignment     = Alignment(horizontal='right', vertical='center')
        c.border        = B(t='medium', r='medium', b='medium', l='thin')
        c.fill          = fill(GOLD)

        # ---- 保存 ----
        tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
        tmp.close()
        wb.save(tmp.name)

        safe_date = invoice_date.replace("年","").replace("月","").replace("日","")
        filename = f"{customer_no}_請求書_{safe_date}.xlsx"
        print(f"\n📄 XLSX請求書生成: {filename}  ({len(cases)}件 合計{total:,}円)")
        self.send_file(tmp.name, filename,
                       "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        os.unlink(tmp.name)

    def handle_add_schedule(self, data):
        """予定を案件表に追加（お客様なしエントリー → Notionカレンダーに同期）"""
        title      = data.get("title","").strip()
        date_str   = data.get("date","")
        memo       = data.get("memo","").strip()
        start_time = data.get("startTime","").strip()  # e.g. "09:00"
        end_time   = data.get("endTime","").strip()    # e.g. "10:00"
        if not title or not date_str:
            self.send_json(400, {"ok": False, "error": "タイトルと日付は必須"}); return

        # 時刻つきの場合はdatetimeフォーマットに変換
        start_dt = f"{date_str}T{start_time}:00+09:00" if start_time else date_str
        end_dt   = f"{date_str}T{end_time}:00+09:00"   if end_time   else None

        date_value = {"start": start_dt}
        if end_dt:
            date_value["end"] = end_dt

        props = {
            "当方案件番号":     {"title": [{"text": {"content": title}}]},
            "案件締切日・進行": {"date": date_value},
            "進捗":            {"status": {"name": "未着手"}},
        }
        if memo:
            props["備考"] = {"rich_text": [{"text": {"content": memo}}]}

        result, err = notion_request("POST", "/pages", {
            "parent": {"database_id": CASE_DB_ID},
            "properties": props,
        })
        if result:
            print(f"  📅 Notion予定追加: 「{title}」{start_dt}")
            self.send_json(200, {"ok": True, "id": result.get("id","")})
        else:
            self.send_json(400, {"ok": False, "error": err})

    def handle_delete_schedule(self, data):
        """予定エントリーをアーカイブ（Notionから削除）"""
        page_id = data.get("id","")
        if not page_id:
            self.send_json(400, {"ok": False, "error": "idが必要"}); return
        # Notion REST API: ページをアーカイブ（= 削除）
        result, err = notion_request("PATCH", f"/pages/{page_id}", {"archived": True})
        if result is not None:
            self.send_json(200, {"ok": True})
        else:
            self.send_json(400, {"ok": False, "error": err})

    def handle_get_all_customers(self):
        """全顧客を取得（顧客情報ツール用）"""
        customers = []
        cursor = None
        while True:
            body = {"sorts": [{"property": "お客様No.", "direction": "ascending"}], "page_size": 100}
            if cursor:
                body["start_cursor"] = cursor
            result, err = notion_request("POST", f"/databases/{CUSTOMER_DB_ID}/query", body)
            if not result:
                self.send_json(400, {"ok": False, "error": err}); return
            for page in result.get("results", []):
                try:
                    props   = page["properties"]
                    no      = (props["お客様No."]["rich_text"] or [{}])[0].get("plain_text","").strip()
                    name    = (props["クライアント名"]["rich_text"] or [{}])[0].get("plain_text","").strip()
                    title_rt= (props["備考"]["title"] or [{}])
                    page_name = title_rt[0].get("plain_text","").strip() if title_rt else ""
                    status  = (props.get("取引状況",{}).get("select") or {}).get("name","")
                    contact = (props.get("お客様優先連絡方法",{}).get("select") or {}).get("name","")
                    kind    = (props.get("種別",{}).get("select") or {}).get("name","")
                    notes_rt= props.get("重要備考",{}).get("rich_text") or []
                    notes   = "".join(t.get("plain_text","") for t in notes_rt)
                    if no or name:
                        customers.append({
                            "id": page["id"], "no": no, "name": name,
                            "pageName": page_name, "status": status,
                            "contact": contact, "kind": kind, "notes": notes,
                        })
                except Exception:
                    pass
            if not result.get("has_more"):
                break
            cursor = result.get("next_cursor")
        print(f"\n👥 全顧客取得: {len(customers)}件")
        self.send_json(200, {"ok": True, "customers": customers})

    def handle_bulk_update_customers(self, data):
        """複数顧客の取引状況を一括変更"""
        ids    = data.get("ids", [])
        status = data.get("status", "")
        if not ids or not status:
            self.send_json(400, {"ok": False, "error": "ids と status が必要"}); return
        ok_list, fail_list = [], []
        for pid in ids:
            result, err = notion_request("PATCH", f"/pages/{pid}", {
                "properties": {"取引状況": {"select": {"name": status}}}
            })
            if result:
                ok_list.append(pid)
            else:
                fail_list.append({"id": pid, "error": err})
        self.send_json(200, {"ok": True, "updated": len(ok_list), "failed": len(fail_list), "errors": fail_list})

    def handle_bulk_archive_customers(self, data):
        """複数顧客をアーカイブ（削除）"""
        ids = data.get("ids", [])
        if not ids:
            self.send_json(400, {"ok": False, "error": "ids が必要"}); return
        ok_list, fail_list = [], []
        for pid in ids:
            result, err = notion_request("PATCH", f"/pages/{pid}", {"archived": True})
            if result:
                ok_list.append(pid)
            else:
                fail_list.append({"id": pid, "error": err})
        self.send_json(200, {"ok": True, "archived": len(ok_list), "failed": len(fail_list)})

    def handle_get_customers(self):
        """取引中・頻度低め・取引開始準備中のお客様をNotionから取得"""
        ACTIVE_STATUSES = {"取引中", "頻度低め", "取引開始準備中"}
        customers = []
        cursor = None
        while True:
            body = {"page_size": 100}
            if cursor:
                body["start_cursor"] = cursor
            result, err = notion_request("POST", f"/databases/{CUSTOMER_DB_ID}/query", body)
            if not result:
                self.send_json(400, {"ok": False, "error": err})
                return
            for page in result.get("results", []):
                try:
                    props = page["properties"]
                    no    = (props["お客様No."]["rich_text"] or [{}])[0].get("plain_text","").strip()
                    name  = (props["クライアント名"]["rich_text"] or [{}])[0].get("plain_text","").strip()
                    status = (props.get("取引状況",{}).get("select") or {}).get("name","")
                    contact = (props.get("お客様優先連絡方法",{}).get("select") or {}).get("name","")
                    notes_rt = props.get("重要備考",{}).get("rich_text") or []
                    notes = "".join(t.get("plain_text","") for t in notes_rt)
                    if no and name and status in ACTIVE_STATUSES:
                        # チャンネル情報をパース
                        channels = []
                        for line in notes.split("\n"):
                            line = line.strip()
                            import re
                            m = re.match(r'^([A-Z]{1,2})案件[：:]\s*(.+)', line)
                            if m:
                                channels.append({"lbl": m.group(1), "name": m.group(2).strip()})
                        customers.append({
                            "no": no, "name": name, "status": status,
                            "contact": contact, "channels": channels,
                            "pageId": page["id"],
                        })
                except Exception:
                    pass
            if not result.get("has_more"):
                break
            cursor = result.get("next_cursor")
        print(f"\n📋 顧客一覧取得: {len(customers)}件")
        self.send_json(200, {"ok": True, "customers": customers})

    def handle_register(self, data):
        print(f"\n📝 案件登録: {data.get('number','?')}  ({data.get('customerNo','?')}様)")

        props = {
            "当方案件番号": {
                "title": [{"text": {"content": data.get("number", "")}}]
            },
            "お客様no/名": {
                "rich_text": [{"text": {"content": data.get("customerNo", "")}}]
            },
            "進捗": {"status": {"name": data.get("progress", "未着手")}},
        }

        if data.get("deadline"):
            props["案件締切日・進行"] = {"date": {"start": data["deadline"]}}
        if data.get("materialName"):
            props["備考/素材名"] = {"rich_text": [{"text": {"content": data["materialName"]}}]}
        if data.get("fileName"):
            props["指定案件ファイル名"] = {"rich_text": [{"text": {"content": data["fileName"]}}]}
        if data.get("memo"):
            props["備考"] = {"rich_text": [{"text": {"content": data["memo"]}}]}
        if data.get("price"):
            try:
                price_val = float(data["price"])
                props["単価"] = {"number": price_val}
                # 外注費がある場合は粗利を計算してテキストで記録
                cost_str = data.get("outsourceCost", "")
                if cost_str:
                    cost_val = float(cost_str)
                    gross = price_val - cost_val
                    def fmt(n):
                        return f"{int(n):,}"
                    gross_text = f"{fmt(price_val)} - {fmt(cost_val)} = {fmt(gross)}"
                    props["粗利（単価-外注費）"] = {"rich_text": [{"text": {"content": gross_text}}]}
            except Exception:
                pass

        result, err = notion_request("POST", "/pages", {
            "parent": {"database_id": CASE_DB_ID},
            "properties": props,
        })

        if result:
            print(f"  ✅ 登録完了: {result.get('url','')}")
            self.send_json(200, {"ok": True, "url": result.get("url", ""), "id": result.get("id", "")})
        else:
            print(f"  ❌ エラー: {err}")
            self.send_json(400, {"ok": False, "error": err})

    def handle_update_customer(self, data):
        no = data.get("customerNo", "")
        entry = data.get("entry", "")
        print(f"\n⭐ 顧客情報追記: {no}番 → {entry}")

        page_id = CUSTOMER_PAGES.get(no)
        if not page_id:
            self.send_json(404, {"ok": False, "error": f"顧客No.{no}のページが見つかりません"})
            return

        # 既存の重要備考を取得
        page, err = notion_request("GET", f"/pages/{page_id}")
        if not page:
            self.send_json(400, {"ok": False, "error": err})
            return

        existing = ""
        try:
            rich = page["properties"]["重要備考"]["rich_text"]
            existing = "".join(t["plain_text"] for t in rich)
        except Exception:
            pass

        new_text = (existing.rstrip() + "\n" + entry).strip()

        result, err = notion_request("PATCH", f"/pages/{page_id}", {
            "properties": {
                "重要備考": {"rich_text": [{"text": {"content": new_text}}]}
            }
        })

        if result:
            print(f"  ✅ 追記完了")
            self.send_json(200, {"ok": True})
        else:
            print(f"  ❌ エラー: {err}")
            self.send_json(400, {"ok": False, "error": err})

    def handle_register_customer(self, data):
        no      = data.get("customerNo", "")
        name    = data.get("name", "")
        print(f"\n👤 顧客登録: {no}番 {name}様")

        props = {
            "備考": {"title": [{"text": {"content": f"{no}_{name}様_補足資料"}}]},
            "お客様No.": {"rich_text": [{"text": {"content": no}}]},
            "クライアント名": {"rich_text": [{"text": {"content": name}}]},
        }
        if data.get("type"):
            props["種別"] = {"select": {"name": data["type"]}}
        if data.get("status"):
            props["取引状況"] = {"select": {"name": data["status"]}}
        if data.get("contact"):
            props["お客様優先連絡方法"] = {"select": {"name": data["contact"]}}
        if data.get("notes"):
            props["重要備考"] = {"rich_text": [{"text": {"content": data["notes"]}}]}

        # ① お客様ページを作成
        cust_result, err = notion_request("POST", "/pages", {
            "parent": {"database_id": CUSTOMER_DB_ID},
            "properties": props,
        })
        if not cust_result:
            print(f"  ❌ 顧客登録エラー: {err}")
            self.send_json(400, {"ok": False, "error": err})
            return

        cust_page_id = cust_result["id"]
        cust_url     = cust_result.get("url", "")
        print(f"  ✅ 顧客ページ作成: {cust_url}")

        # ② 補足資料ページを作成（子ページ）
        sub_title = f"{no}_{name}様_補足資料"
        sub_result, sub_err = notion_request("POST", "/pages", {
            "parent": {"page_id": cust_page_id},
            "properties": {
                "title": {"title": [{"text": {"content": sub_title}}]}
            },
            "children": [
                {"object": "block", "type": "heading_2",
                 "heading_2": {"rich_text": [{"text": {"content": "基本情報"}}]}},
                {"object": "block", "type": "paragraph",
                 "paragraph": {"rich_text": [{"text": {"content": f"お客様No.: {no}\nクライアント名: {name}"}}]}},
                {"object": "block", "type": "heading_2",
                 "heading_2": {"rich_text": [{"text": {"content": "案件一覧"}}]}},
                {"object": "block", "type": "paragraph",
                 "paragraph": {"rich_text": [{"text": {"content": "（案件を追記してください）"}}]}},
                {"object": "block", "type": "heading_2",
                 "heading_2": {"rich_text": [{"text": {"content": "連絡先・重要事項"}}]}},
                {"object": "block", "type": "paragraph",
                 "paragraph": {"rich_text": [{"text": {"content": data.get("notes", "（重要事項を追記してください）")}}]}},
            ]
        })

        sub_url = sub_result.get("url", "") if sub_result else ""
        if sub_result:
            print(f"  ✅ 補足資料ページ作成: {sub_url}")
        else:
            print(f"  ⚠️  補足資料ページ作成失敗: {sub_err}")

        self.send_json(200, {
            "ok": True,
            "customerUrl": cust_url,
            "subPageUrl": sub_url,
            "customerNo": no,
            "name": name,
        })


    def handle_get_invoice_data(self):
        """月別案件データをNotionから取得してお客様ごとにグループ化"""
        import re as _re
        from urllib.parse import urlparse, parse_qs
        from datetime import date

        qs = parse_qs(urlparse(self.path).query)
        month_str = qs.get("month", [None])[0]

        today = date.today()
        if month_str:
            try:
                year  = int(month_str.split("-")[0])
                month = int(month_str.split("-")[1])
            except Exception:
                year, month = today.year, today.month
        else:
            # デフォルト: 当月
            year, month = today.year, today.month

        start = f"{year}-{month:02d}-01"
        end   = f"{year+1}-01-01" if month == 12 else f"{year}-{month+1:02d}-01"
        print(f"\n📑 請求データ取得: {year}年{month}月 ({start} 〜 {end})")

        query_body = {
            "filter": {
                "and": [
                    {"property": "案件締切日・進行", "date": {"on_or_after": start}},
                    {"property": "案件締切日・進行", "date": {"before": end}},
                ]
            },
            "sorts": [{"property": "お客様no/名", "direction": "ascending"}],
            "page_size": 100,
        }

        cases = []
        cursor = None
        while True:
            if cursor:
                query_body["start_cursor"] = cursor
            result, err = notion_request("POST", f"/databases/{CASE_DB_ID}/query", query_body)
            if not result:
                self.send_json(400, {"ok": False, "error": err})
                return
            for page in result.get("results", []):
                try:
                    props   = page["properties"]
                    number  = (props["当方案件番号"]["title"] or [{}])[0].get("plain_text","").strip()
                    customer= (props["お客様no/名"]["rich_text"] or [{}])[0].get("plain_text","").strip()
                    price_v = props["単価"]["number"]
                    price   = price_v if price_v is not None else 0
                    note    = (props["備考/素材名"]["rich_text"] or [{}])[0].get("plain_text","").strip()
                    dl      = (props["案件締切日・進行"]["date"] or {}).get("start","")
                    status  = (props["進捗"]["status"] or {}).get("name","")
                    if number and customer:
                        cases.append({"number": number, "customer": customer,
                                      "price": price, "note": note,
                                      "date": dl, "status": status})
                except Exception as e:
                    pass
            if not result.get("has_more"):
                break
            cursor = result.get("next_cursor")

        # お客様ごとにグループ化
        groups = {}
        for c in cases:
            key = c["customer"]
            if key not in groups:
                groups[key] = {"customer": key, "cases": [], "total": 0}
            groups[key]["cases"].append(c)
            groups[key]["total"] += c["price"]

        print(f"  → {len(cases)}件 / {len(groups)}お客様")
        self.send_json(200, {
            "ok": True,
            "month": f"{year}-{month:02d}",
            "label": f"{year}年{month}月",
            "customers": list(groups.values()),
        })

    def find_customer_page_id(self, customer_no):
        """お客様NoからNotionページIDを動的検索"""
        if customer_no in CUSTOMER_PAGES:
            return CUSTOMER_PAGES[customer_no]
        body = {
            "filter": {"property": "お客様No.", "rich_text": {"equals": customer_no}},
            "page_size": 1,
        }
        result, _ = notion_request("POST", f"/databases/{CUSTOMER_DB_ID}/query", body)
        if result and result.get("results"):
            pid = result["results"][0]["id"]
            CUSTOMER_PAGES[customer_no] = pid
            return pid
        return None

    def find_invoice_storage_page(self, customer_no):
        """お客様Noの請求書格納庫ページをNotionで検索"""
        # キャッシュキー
        cache_key = f"invoice_{customer_no}"
        if cache_key in CUSTOMER_PAGES:
            return CUSTOMER_PAGES[cache_key]
        result, _ = notion_request("POST", "/search", {
            "query": f"{customer_no}様_請求書格納庫",
            "filter": {"value": "page", "property": "object"},
            "page_size": 10,
        })
        if result:
            for page in result.get("results", []):
                title_rt = (page.get("properties", {}).get("title", {}).get("title") or [])
                title = title_rt[0].get("plain_text", "") if title_rt else ""
                if f"{customer_no}様" in title and "請求書格納庫" in title:
                    pid = page["id"]
                    CUSTOMER_PAGES[cache_key] = pid
                    print(f"  🔍 格納庫発見: {title} ({pid})")
                    return pid
        return None

    def handle_record_invoice(self, data):
        """請求記録をお客様の請求書格納庫ページにトグルで追記"""
        records  = data.get("records", [])
        inv_date = data.get("invoiceDate", "")
        results  = []
        for rec in records:
            cno     = rec.get("customerNo", "")
            amount  = rec.get("amount", 0)
            month   = rec.get("month", "")   # 例: "2026年3月"
            numbers = rec.get("caseNumbers", [])

            # "2026年3月" → "2026年_3月" (格納庫の命名規則に合わせる)
            toggle_title = month.replace("年", "年_") if "年" in month else month

            page_id = self.find_invoice_storage_page(cno)
            if not page_id:
                print(f"  ⚠️  請求書格納庫未発見: {cno}様")
                results.append({"customerNo": cno, "ok": False,
                                 "error": f"{cno}様_請求書格納庫 ページが見つかりません"})
                continue

            detail_text = (f"請求日: {inv_date}　"
                           f"金額: ¥{int(amount):,}　"
                           f"案件 {len(numbers)}件: {', '.join(numbers)}")
            block = {
                "children": [{
                    "object": "block",
                    "type": "toggle",
                    "toggle": {
                        "rich_text": [{"type": "text", "text": {"content": toggle_title}}],
                        "children": [{
                            "object": "block",
                            "type": "paragraph",
                            "paragraph": {
                                "rich_text": [{"type": "text",
                                               "text": {"content": detail_text}}]
                            }
                        }]
                    }
                }]
            }
            res, err = notion_request("PATCH", f"/blocks/{page_id}/children", block)
            if res:
                print(f"  ✅ 格納庫に追記: {cno} → {toggle_title} {detail_text[:50]}")
                results.append({"customerNo": cno, "ok": True, "page": page_id})
            else:
                print(f"  ❌ 追記失敗: {cno} → {err}")
                results.append({"customerNo": cno, "ok": False, "error": err})

        self.send_json(200, {"ok": True, "results": results})


if __name__ == "__main__":
    if NOTION_TOKEN == "secret_ここに貼り付け":
        print("=" * 55)
        print("⚠️  Notionトークンが設定されていません")
        print()
        print("1. https://www.notion.so/my-integrations を開く")
        print("2. 「新しいインテグレーション」をクリック")
        print("3. 名前: 案件登録ツール → 送信")
        print("4. 「シークレット」のトークンをコピー")
        print("5. このファイル（notion_server.py）の")
        print("   NOTION_TOKEN = の行に貼り付けて保存")
        print("=" * 55)
        sys.exit(1)

    import socket
    local_ip = socket.gethostbyname(socket.gethostname())

    print("=" * 55)
    print("📁 案件登録サーバー起動中...")
    print(f"   PC:     http://localhost:{PORT}")
    print(f"   iPhone: http://{local_ip}:{PORT}")
    print("   （同じWiFiで接続してください）")
    print("   停止: Ctrl+C")
    print("=" * 55)

    server = HTTPServer(("0.0.0.0", PORT), Handler)
    try:
        server.serve_forever()
    except KeyboardInterrupt:
        print("\n✅ サーバーを停止しました")
